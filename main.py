# -*- coding: utf-8 -*-
"""
基金净值自动更新器

依赖安装（在命令行中运行）：
    pip install akshare pandas openpyxl

功能概述：
1. 自动更新三个 Sheet："主推公募"、"ETF"、"个人关注产品"（10列表头 + 数据更新）
2. "私募资管" Sheet 完全不处理，只完整复制继承（支持用户随意自定义表头、加列、改格式）
3. 历史文件继承（跨日自动找最近历史文件复制所有内容）
4. 全量基金名称映射（解决建仓期名称问题）
5. 所有收益率均使用累计净值计算（含分红复利）
6. 近1年波动率（年化标准差×√252）和夏普比率（无风险利率1.5%，无额外请求）
7. 每次运行后按当日收益率对各Sheet降序排列（None置底）
8. 更新时间统一写在 K1
9. 列宽自动调整（支持长中文基金名称完整显示）
10. 数据获取速度优化（0.3-0.6秒随机间隔）
"""

import os
import sys
import time
import logging
import math
import re
import shutil
import random
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# ===================== 网络请求补丁 =====================
# 修复 AKShare 内部 requests.get 在部分网络环境（如存在无响应的 CDN 节点或 IPv6 丢包）下
# 因默认无 timeout 而导致长时间阻塞（Windows TCP timeout 约 42s）的问题。
# 强制加入连接超时机制（3秒），使其能极速 Failover（故障转移）到下一个健康的 IP。
import requests
_original_get = requests.get

def _patched_get(*args, **kwargs):
    if 'timeout' not in kwargs:
        kwargs['timeout'] = (3.0, 10.0)
    return _original_get(*args, **kwargs)

requests.get = _patched_get

# ===================== 依赖检查 =====================

try:
    import akshare as ak  # type: ignore
    HAS_AKSHARE = True
except ImportError:
    HAS_AKSHARE = False

# ===================== 常量配置 =====================

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "fund_data"

# Sheet 配置
AUTO_SHEETS = ["主推公募", "ETF", "个人关注产品"]  # 自动处理的Sheet
MANUAL_SHEET = "私募资管"  # 手动Sheet，不自动处理
SHEET_ORDER = ["主推公募", "ETF", "私募资管", "个人关注产品"]  # Sheet显示顺序
ALL_SHEETS = AUTO_SHEETS + [MANUAL_SHEET]  # 所有需要存在的Sheet

# 表头定义（10列，A-J）
HEADERS = [
    "产品代码", "产品名称", "最新单位净值",
    "当日收益率(%)", "近7天收益率(%)", "近1月收益率(%)", "近1年收益率(%)",
    "近1年波动率(%)", "近1年夏普",
    "成立以来收益率(%)"
]

# 更新时间配置（K1，J列已用于成立以来收益率）
TIMESTAMP_CELL = "K1"
TIMESTAMP_PREFIX = "更新时间："

# 年化无风险利率（用于夏普比率计算，参考中国1年期国债收益率）
RISK_FREE_RATE_ANNUAL = 0.015

# 请求节流配置（随机间隔，避免触发风控）
REQUEST_INTERVAL_MIN = 0.3
REQUEST_INTERVAL_MAX = 0.6

# ===================== 日志配置 =====================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)

# ===================== 全局变量 =====================

# 全量基金名称映射缓存（在main函数开头一次性加载）
_fund_name_map = None

# ===================== 文件操作函数 =====================

def ensure_data_dir():
    """确保数据目录存在"""
    if not DATA_DIR.exists():
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        logging.info(f"已创建数据目录：{DATA_DIR}")


def get_today_filename():
    """生成当天 Excel 文件名：fund_tracker_YYYY-MM-DD.xlsx"""
    today_str = datetime.now().strftime("%Y-%m-%d")
    return DATA_DIR / f"fund_tracker_{today_str}.xlsx"


def find_latest_historical_file(exclude_date: datetime = None):
    """
    查找最近的历史文件（按日期倒序）
    
    参数:
        exclude_date: 要排除的日期（通常是当天），如果为None则使用当天
    返回:
        最近的历史文件路径（Path对象），如果没找到则返回None
    """
    if exclude_date is None:
        exclude_date = datetime.now()
    
    ensure_data_dir()
    today_date = exclude_date.date()
    pattern = re.compile(r'^fund_tracker_(\d{4}-\d{2}-\d{2})\.xlsx$')
    historical_files = []
    
    try:
        for filename in os.listdir(DATA_DIR):
            match = pattern.match(filename)
            if match:
                try:
                    file_date_str = match.group(1)
                    file_date = datetime.strptime(file_date_str, "%Y-%m-%d").date()
                    file_path = DATA_DIR / filename
                    if file_date < today_date and file_path.exists():
                        historical_files.append((file_date, file_path))
                except ValueError:
                    continue
    except OSError:
        return None
    
    if not historical_files:
        return None
    
    historical_files.sort(key=lambda x: x[0], reverse=True)
    return historical_files[0][1]


def copy_workbook(src_path: Path, dst_path: Path):
    """
    复制工作簿：将所有Sheet和所有数据完整复制到新文件，保留格式
    
    参数:
        src_path: 源文件路径
        dst_path: 目标文件路径
    返回:
        复制后的工作簿对象
    """
    shutil.copy2(src_path, dst_path)
    return load_workbook(dst_path)


# ===================== 工作簿创建与加载 =====================

def create_template_workbook(file_path: Path):
    """创建带有所有 Sheet 和表头的模板 Excel 文件，按 SHEET_ORDER 顺序创建"""
    wb = Workbook()
    wb.remove(wb.active)
    
    for sheet_name in SHEET_ORDER:
        ws = wb.create_sheet(title=sheet_name)
        # 私募资管不写入表头，保留手动定制；其他使用完整表头
        if sheet_name != MANUAL_SHEET:
            ws.append(HEADERS)
    
    wb.save(file_path)
    logging.info(f"已创建当日模板文件：{file_path}")
    return wb


def ensure_sheet_headers(ws):
    """
    确保 Sheet 的表头完整，如果列数不足则添加新列
    
    返回:
        True 如果表头已更新，False 如果表头已完整
    """
    # 私募资管Sheet不自动处理表头，保留手动内容
    if ws.title == MANUAL_SHEET:
        return False
    
    max_col = ws.max_column
    header_count = len(HEADERS)
    
    if max_col < header_count:
        logging.info(f"Sheet '{ws.title}' 表头列数不足（现有{max_col}列，需要{header_count}列），正在添加新列...")
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)
        return True
    else:
        # 检查表头内容是否正确
        headers_match = True
        for col_idx, expected_header in enumerate(HEADERS, start=1):
            current_value = ws.cell(row=1, column=col_idx).value
            if str(current_value) != str(expected_header):
                headers_match = False
                break
        if not headers_match:
            logging.info(f"Sheet '{ws.title}' 表头内容不一致，正在更新...")
            for col_idx, header in enumerate(HEADERS, start=1):
                ws.cell(row=1, column=col_idx, value=header)
            return True
    return False


def reorder_sheets(wb):
    """按照 SHEET_ORDER 顺序重新排列工作簿中的 Sheet"""
    sheet_dict = {sheet.title: sheet for sheet in wb.worksheets}
    ordered_sheets = []
    for sheet_name in SHEET_ORDER:
        if sheet_name in sheet_dict:
            ordered_sheets.append(sheet_dict[sheet_name])
    wb._sheets = ordered_sheets


def merge_legacy_sheets(wb):
    """
    合并历史文件中的"私募"和"资管"Sheet到"私募资管"Sheet（兼容旧版本）
    如果历史文件已有"私募资管"，直接保留；否则合并"私募"和"资管"的数据
    """
    if MANUAL_SHEET in wb.sheetnames:
        logging.info(f"历史文件已有'{MANUAL_SHEET}'Sheet，直接保留")
        return
    
    has_private = "私募" in wb.sheetnames
    has_asset = "资管" in wb.sheetnames
    
    if not has_private and not has_asset:
        ws_new = wb.create_sheet(title=MANUAL_SHEET)
        logging.info(f"已创建新的'{MANUAL_SHEET}'Sheet（保留手动定制）")
        return
    
    # 创建"私募资管"Sheet，复制原表头而不是写入固定表头
    ws_merged = wb.create_sheet(title=MANUAL_SHEET)
    
    # 从第一个有数据的Sheet复制表头（如果有）
    if has_private:
        ws_private = wb["私募"]
        if ws_private.max_row >= 1:
            for col_idx in range(1, ws_private.max_column + 1):
                header_value = ws_private.cell(row=1, column=col_idx).value
                if header_value:
                    ws_merged.cell(row=1, column=col_idx, value=header_value)
    elif has_asset:
        ws_asset = wb["资管"]
        if ws_asset.max_row >= 1:
            for col_idx in range(1, ws_asset.max_column + 1):
                header_value = ws_asset.cell(row=1, column=col_idx).value
                if header_value:
                    ws_merged.cell(row=1, column=col_idx, value=header_value)
    
    # 合并数据
    seen_codes = set()
    row_count = 1
    
    for old_sheet_name in ["私募", "资管"]:
        if old_sheet_name not in wb.sheetnames:
            continue
        
        ws_old = wb[old_sheet_name]
        logging.info(f"正在合并'{old_sheet_name}'Sheet的数据（共{ws_old.max_row}行）")
        
        for row_idx in range(2, ws_old.max_row + 1):
            code = ws_old.cell(row=row_idx, column=1).value
            if code:
                code_str = str(code).strip()
                if code_str and code_str not in seen_codes:
                    seen_codes.add(code_str)
                    row_count += 1
                    # 复制数据：A产品代码、B产品名称、C最新净值、F近1月、G近1年、H成立以来
                    ws_merged.cell(row=row_count, column=1, value=ws_old.cell(row=row_idx, column=1).value)
                    ws_merged.cell(row=row_count, column=2, value=ws_old.cell(row=row_idx, column=2).value)
                    ws_merged.cell(row=row_count, column=3, value=ws_old.cell(row=row_idx, column=3).value)
                    # 检查旧Sheet是否有F列和G列（近1月、近1年）
                    if ws_old.max_column >= 6:
                        ws_merged.cell(row=row_count, column=4, value=ws_old.cell(row=row_idx, column=6).value)
                    if ws_old.max_column >= 7:
                        ws_merged.cell(row=row_count, column=5, value=ws_old.cell(row=row_idx, column=7).value)
                    # 检查旧Sheet是否有H列（成立以来），如果有则复制到新Sheet的F列
                    if ws_old.max_column >= 8:
                        ws_merged.cell(row=row_count, column=6, value=ws_old.cell(row=row_idx, column=8).value)
    
    logging.info(f"已合并'私募'和'资管'Sheet到'{MANUAL_SHEET}'，共{row_count-1}条数据")
    
    # 删除旧的"私募"和"资管"Sheet
    if has_private:
        wb.remove(wb["私募"])
        logging.info("已删除旧的'私募'Sheet")
    if has_asset:
        wb.remove(wb["资管"])
        logging.info("已删除旧的'资管'Sheet")


def load_or_inherit_workbook():
    """
    加载或创建当天的工作簿。如果当天文件不存在，从最近的历史文件继承
    
    返回:
        (工作簿对象, 文件路径)
    """
    ensure_data_dir()
    file_path = get_today_filename()
    
    if file_path.exists():
        # data_only=True: 确保读取纯数据，避免 openpyxl 对现有单元格的缓存数据型存留
        wb = load_workbook(file_path, data_only=True)
        logging.info(f"已加载当日文件：{file_path}")

        
        # 确保所有需要的 Sheet 存在且有表头
        for sheet_name in ALL_SHEETS:
            if sheet_name not in wb.sheetnames:
                ws_new = wb.create_sheet(title=sheet_name)
                if sheet_name != MANUAL_SHEET:
                    ws_new.append(HEADERS)
                logging.info(f"已补充缺失的 Sheet：{sheet_name}")
            else:
                ws = wb[sheet_name]
                # 私募资管不处理表头，保留手动内容
                if sheet_name == MANUAL_SHEET:
                    pass
                elif ws.max_row < 1:
                    ws.append(HEADERS)
                else:
                    ensure_sheet_headers(ws)
        
        reorder_sheets(wb)
        return wb, file_path
    else:
        # 当天文件不存在，尝试从历史文件继承
        prev_file = find_latest_historical_file()
        if prev_file:
            logging.info(f"当天文件不存在，从历史文件 {prev_file.name} 继承追踪列表")
            try:
                wb = copy_workbook(prev_file, file_path)
                # 处理"私募"和"资管"Sheet的合并（兼容旧版本）
                merge_legacy_sheets(wb)
                
                # 确保所有需要的 Sheet 存在
                for sheet_name in ALL_SHEETS:
                    if sheet_name not in wb.sheetnames:
                        ws_new = wb.create_sheet(title=sheet_name)
                        if sheet_name != MANUAL_SHEET:
                            ws_new.append(HEADERS)
                        logging.info(f"已补充缺失的 Sheet：{sheet_name}")
                    else:
                        # 私募资管不处理表头，保留手动内容
                        if sheet_name != MANUAL_SHEET:
                            ensure_sheet_headers(wb[sheet_name])
                
                reorder_sheets(wb)
                return wb, file_path
            except Exception as e:
                logging.warning(f"复制历史文件失败：{e}，将创建新模板")
                wb = create_template_workbook(file_path)
                return wb, file_path
        else:
            # 没有历史文件，创建新模板
            logging.info("未找到历史文件，创建新模板")
            wb = create_template_workbook(file_path)
            return wb, file_path


# ===================== 基金代码处理 =====================

def normalize_fund_code(raw):
    """把单元格中的基金代码标准化为6位字符串"""
    if raw is None:
        return None
    try:
        if isinstance(raw, (int, float)):
            return f"{int(raw):06d}"
        s = str(raw).strip()
        if not s:
            return None
        s = "".join(ch for ch in s if ch.isdigit())
        if not s:
            return None
        if len(s) == 6:
            return s
        if len(s) > 6:
            return s[-6:]
        return s.zfill(6)
    except Exception:
        return None


def get_codes_from_sheet(ws):
    """
    从 Sheet 的 A 列（从第二行开始）读取所有非空的产品代码
    
    返回:
        列表：[(row_index, code), ...]
    """
    codes = []
    for row in range(2, ws.max_row + 1):
        raw_value = ws.cell(row=row, column=1).value
        code = normalize_fund_code(raw_value)
        if code:
            codes.append((row, code))
    return codes


# ===================== 基金名称映射 =====================

def load_fund_name_map():
    """
    一次性加载全量基金名称映射
    
    返回:
        字典：{基金代码: 基金名称}
    """
    global _fund_name_map
    if _fund_name_map is not None:
        return _fund_name_map
    
    _fund_name_map = {}
    if not HAS_AKSHARE:
        return _fund_name_map
    
    try:
        all_funds = ak.fund_name_em()
        if all_funds is not None and not all_funds.empty:
            code_col = None
            name_col = None
            for col in all_funds.columns:
                col_str = str(col)
                if "代码" in col_str and code_col is None:
                    code_col = col
                if ("简称" in col_str or "名称" in col_str or "name" in col_str.lower()) and name_col is None:
                    name_col = col
            
            if code_col and name_col:
                for _, row in all_funds.iterrows():
                    try:
                        fund_code = str(row[code_col]).strip()
                        fund_name = str(row[name_col]).strip()
                        if fund_code and fund_name:
                            _fund_name_map[fund_code] = fund_name
                    except Exception:
                        continue
    except Exception as e:
        logging.warning(f"获取全量基金名称映射失败：{e}，将尝试备用接口")
        try:
            name_df = ak.fund_em_fund_name()
            if name_df is not None and not name_df.empty:
                code_col = None
                name_col = None
                for col in name_df.columns:
                    col_str = str(col)
                    if "代码" in col_str:
                        code_col = col
                    if "简称" in col_str or "名称" in col_str:
                        name_col = col
                
                if code_col and name_col:
                    for _, row in name_df.iterrows():
                        try:
                            fund_code = str(row[code_col]).strip()
                            fund_name = str(row[name_col]).strip()
                            if fund_code and fund_name:
                                _fund_name_map[fund_code] = fund_name
                        except Exception:
                            continue
        except Exception:
            pass
    
    return _fund_name_map


def fetch_fund_name_from_akshare(code: str):
    """
    尝试从 AKShare 获取基金名称（用于建仓期基金）
    优先使用已加载的全量映射，如果不存在则返回None，避免重复请求导致卡顿和封禁
    
    返回:
        基金名称字符串，如果失败则返回 None
    """
    global _fund_name_map
    if _fund_name_map and code in _fund_name_map:
        return _fund_name_map[code]
        
    return None


# ===================== 基金数据获取 =====================

def fetch_fund_data(code: str):
    """
    获取基金数据（使用单位净值走势接口，重构复权净值）
    所有收益率均基于日增长率的几何累乘计算（反映真实的红利再投资复利曲线）

    返回字典：
        {
            "name": 名称或 None,
            "nav": 最新单位净值(float) 或 None,
            "today_pct": 当日收益率(float, %) 或 None,
            "week_pct": 近7天收益率(float, %) 或 None,
            "month_pct": 近1月收益率(float, %) 或 None,
            "year_pct": 近1年收益率(float, %) 或 None,
            "since_inception_pct": 成立以来收益率(float, %) 或 None,
            "is_building_period": True/False (是否为建仓期/封闭期基金)
        }
    """
    if not HAS_AKSHARE:
        raise RuntimeError("akshare 未安装，无法获取基金数据")

    # 使用单位净值走势作为数据源（其日增长率是除权平滑后的真实涨跌）
    try:
        df = ak.fund_open_fund_info_em(symbol=code, indicator="单位净值走势")
    except Exception as e:
        name = fetch_fund_name_from_akshare(code)
        if name:
            raise RuntimeError(f"建仓期/封闭期基金: {e}")
        else:
            raise RuntimeError(f"AKShare 获取单位净值走势失败：{e}")

    # 检查数据是否为空（建仓期基金）
    if df is None or df.empty:
        name = fetch_fund_name_from_akshare(code)
        if name:
            return {
                "name": name,
                "nav": None,
                "today_pct": None,
                "week_pct": None,
                "month_pct": None,
                "year_pct": None,
                "since_inception_pct": None,
                "is_building_period": True,
            }
        else:
            raise RuntimeError("AKShare 返回空数据且无法获取基金名称")

    # 规范字段名
    col_map = {}
    for col in df.columns:
        c = str(col)
        if "净值日期" in c or "日期" in c or "date" in c.lower():
            col_map["date"] = col
        elif "单位净值" in c and "累计" not in c:
            col_map["unit_nav"] = col
        elif "日增长率" in c or "涨跌幅" in c:
            col_map["daily_growth"] = col
        elif "基金名称" in c or "名称" in c or "name" in c.lower():
            col_map["name"] = col

    if "date" not in col_map or "unit_nav" not in col_map or "daily_growth" not in col_map:
        raise RuntimeError("AKShare 单位净值数据列结构无法解析")

    df = df.copy()
    df[col_map["date"]] = pd.to_datetime(df[col_map["date"]], errors="coerce")
    df = df.dropna(subset=[col_map["date"]])
    df = df.sort_values(col_map["date"])

    if len(df) == 0:
        name = fetch_fund_name_from_akshare(code)
        if name:
            return {
                "name": name,
                "nav": None,
                "today_pct": None,
                "week_pct": None,
                "month_pct": None,
                "year_pct": None,
                "since_inception_pct": None,
                "is_building_period": True,
            }
        else:
            raise RuntimeError("AKShare 处理后数据为空且无法获取基金名称")

    # 核心复利重构：清洗日增长率并计算虚拟复权净值曲线
    daily_str = df[col_map["daily_growth"]].astype(str).str.replace("%", "", regex=False)
    # 将无法转换为数字的置为0（比如上市第一天的空缺或节假日）
    daily_float = pd.to_numeric(daily_str, errors="coerce").fillna(0.0) / 100.0
    
    # 构建从1.0开始的几何累乘真实复权曲线
    df["adj_nav"] = (1.0 + daily_float).cumprod()

    latest = df.iloc[-1]

    # 获取最新单位净值 (C列)
    try:
        nav = float(latest[col_map["unit_nav"]])
    except Exception:
        nav = None

    # 获取基金名称
    name = None
    if "name" in col_map:
        try:
            name = str(latest[col_map["name"]])
        except Exception:
            pass
    if not name:
        name = fetch_fund_name_from_akshare(code)

    # 当日收益率：直接取最新的日增长率乘 100
    today_pct = None
    if len(daily_float) > 0:
        today_pct = float(daily_float.iloc[-1]) * 100.0

    # 计算近7天收益率：基于复权净值
    week_pct = None
    if len(df) >= 7:
        try:
            adj_now = float(latest["adj_nav"])
            adj_7_ago = float(df.iloc[-7]["adj_nav"])
            if adj_7_ago > 0:
                week_pct = (adj_now / adj_7_ago - 1.0) * 100.0
        except Exception:
            week_pct = None

    # 计算近1月收益率：30天前的复权净值
    month_pct = None
    if len(df) > 0:
        try:
            latest_date = pd.to_datetime(latest[col_map["date"]])
            target_date = latest_date - pd.Timedelta(days=30)
            mask = df[col_map["date"]] <= target_date
            if mask.any():
                month_df = df[mask]
                if len(month_df) > 0:
                    adj_month_ago = float(month_df.iloc[-1]["adj_nav"])
                    if adj_month_ago > 0:
                        adj_now = float(latest["adj_nav"])
                        month_pct = (adj_now / adj_month_ago - 1.0) * 100.0
        except Exception:
            month_pct = None

    # 计算近1年收益率、波动率、夏普比率
    year_pct = None
    volatility_1y_pct = None
    sharpe_1y = None
    if len(df) > 0:
        try:
            latest_date = pd.to_datetime(latest[col_map["date"]])
            target_date = latest_date - pd.Timedelta(days=365)
            mask = df[col_map["date"]] <= target_date

            if mask.any():
                year_df = df[mask]
                if len(year_df) > 0:
                    adj_year_ago = float(year_df.iloc[-1]["adj_nav"])
                    if adj_year_ago > 0:
                        adj_now = float(latest["adj_nav"])
                        year_pct = (adj_now / adj_year_ago - 1.0) * 100.0

                # 波动率计算：直接利用每日真实的日增长率 (daily_float) 的后一年的部分
                year_mask = df[col_map["date"]] > target_date
                year_data = df[year_mask]
                
                if len(year_data) >= 5:
                    try:
                        # 日收益率序列 (直接基于日增长率，而不是价格相比，这样最真实)
                        year_daily_returns = daily_float[year_mask]
                        if len(year_daily_returns) >= 2:
                            vol_decimal = year_daily_returns.std() * math.sqrt(252)
                            volatility_1y_pct = vol_decimal * 100.0
                            
                            if year_pct is not None and year_pct >= 0 and vol_decimal > 0:
                                sharpe_1y = (year_pct / 100.0 - RISK_FREE_RATE_ANNUAL) / vol_decimal
                            else:
                                sharpe_1y = None
                    except Exception:
                        volatility_1y_pct = None
                        sharpe_1y = None
        except Exception:
            year_pct = None

    # 成立以来收益率：基于复权净值
    since_inception_pct = None
    if len(df) > 0:
        try:
            adj_start = float(df.iloc[0]["adj_nav"])
            adj_now = float(latest["adj_nav"])
            if adj_start > 0:
                since_inception_pct = (adj_now / adj_start - 1.0) * 100.0
                logging.info(f"代码 {code} 使用日增长率复权净值建模完成")
        except Exception:
            since_inception_pct = None

    return {
        "name": name,
        "nav": nav,
        "today_pct": today_pct,
        "week_pct": week_pct,
        "month_pct": month_pct,
        "year_pct": year_pct,
        "volatility_1y_pct": volatility_1y_pct,
        "sharpe_1y": sharpe_1y,
        "since_inception_pct": since_inception_pct,
        "is_building_period": False,
    }



# ===================== Excel格式化函数 =====================


def format_sheet(ws):
    """
    格式化 Sheet（不包含列宽调整）：
    - A列（产品代码）设置为文本格式，防止前导零丢失
    - C列（最新累计净值）：5位小数
    - D-H列（各收益率+波动率）：百分比格式，2位小数
    - I列（夏普比率）：普通数字格式，2位小数（不带%）
    - J列（成立以来收益率）：百分比格式，2位小数
    私募资管Sheet不格式化，保留手动设置
    """
    if ws.title == MANUAL_SHEET:
        return

    # 设置A列为文本格式（从第2行开始）
    if ws.max_row >= 2:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value is not None:
                cell.number_format = "@"

    # 设置数字列格式
    if ws.max_row >= 2:
        # C列：累计净值，5位小数
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=3)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00000"

        # D-H列（当日/7天/1月/1年收益率 + 近1年波动率）：百分比，2位小数
        for col_idx in [4, 5, 6, 7, 8]:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00%"

        # I列（近1年夏普）：普通数字，2位小数，不带%
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=9)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

        # J列（成立以来收益率）：百分比，2位小数
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=10)
            if cell.value is not None and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"



def adjust_column_widths(ws):
    """
    自动调整 Sheet 的所有列宽（在所有数据填充完毕后调用）
    优化支持长中文基金名称完整显示（如"易方达中证香港证券投资主题ETF"）
    私募资管Sheet不调整列宽，保留手动设置
    """
    if ws.title == MANUAL_SHEET:
        return
    
    def calculate_display_width(text):
        """计算文本的显示宽度，中文字符按1.8倍计算"""
        if not text:
            return 0
        width = 0
        for char in str(text):
            if '\u4e00' <= char <= '\u9fff' or '\u3000' <= char <= '\u303f' or '\uff00' <= char <= '\uffef':
                width += 1.8  # 中文字符宽度系数
            else:
                width += 1.0  # 英文字符和数字
        return width
    
    for col in ws.columns:
        max_display_width = 0
        col_letter = get_column_letter(col[0].column)
        
        # 遍历该列的所有单元格（包括表头）
        for cell in col:
            try:
                if cell.value is not None:
                    if isinstance(cell.value, (int, float)):
                        # 判断是否为百分比格式列
                        # H=波动率(%), D/E/F/G=收益率(%), J=成立以来(%); I=夏普(纯数字)
                        is_pct_col = col_letter in ["D", "E", "F", "G", "H", "J"]
                        if is_pct_col:
                            display_value = cell.value * 100
                            display_text = f"{display_value:.2f}%"
                        else:
                            display_text = f"{cell.value:.5f}"
                        width = calculate_display_width(display_text)
                        max_display_width = max(max_display_width, width)
                    else:
                        width = calculate_display_width(str(cell.value))
                        max_display_width = max(max_display_width, width)
            except Exception:
                pass
        
        # 根据列类型设置不同的宽度公式（进一步收紧 A/B 列）
        if col_letter == "A":
            # 产品代码列：更窄，避免空白
            adjusted_width = max_display_width * 1.0 + 1.5
            adjusted_width = min(max(adjusted_width, 8), 25)
        elif col_letter == "B":
            # 产品名称列：再收紧，仍预留约1个中文字符余量
            adjusted_width = max_display_width * 1.10 + 1.5
            adjusted_width = min(max(adjusted_width, 10), 35)
        else:
            # 其他列：保持适中宽度
            adjusted_width = max_display_width * 1.35 + 2
            adjusted_width = min(max(adjusted_width, 10), 38)
        
        ws.column_dimensions[col_letter].width = adjusted_width


def write_update_time(ws):
    """
    在 Sheet 指定单元格写入更新时间（K1）
    清除可能存在的其他位置的更新时间，确保只写入一次
    私募资管Sheet不写入更新时间，保留手动内容
    """
    if ws.title == MANUAL_SHEET:
        return

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    # 清除其他位置可能存在的旧时间戳
    for cell_ref in ["J1", "I1", "H1", "G1"]:
        try:
            if ws[cell_ref].value and str(ws[cell_ref].value).startswith(TIMESTAMP_PREFIX):
                ws[cell_ref].value = None
        except Exception:
            pass
    # 写入到 K1 位置
    ws[TIMESTAMP_CELL] = f"{TIMESTAMP_PREFIX}{now_str}"



# ===================== Sheet排序函数 =====================

def sort_sheet_by_daily_return(ws):
    """
    按当日收益率（D列）对 Sheet 的数据行降序排序
    - None 值（无数据/建仓期）置于底部
    - 私募资管 Sheet 不进行排序
    """
    if ws.title == MANUAL_SHEET:
        return
    if ws.max_row <= 2:  # 只有表头行或无数据，跳过
        return

    total_cols = ws.max_column
    # 读取所有数据行（第2行起）为列表
    rows_data = []
    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, total_cols + 1)]
        if row_values[0] is None or row_values[0] == "":
            rows_data.append((None, row_values))
        else:
            # 第4列（D列，当日收益率）作为排序键
            sort_key = row_values[3] if len(row_values) > 3 else None
            # 空字符串视为无值
            if sort_key == "" or sort_key is None:
                sort_key = None
            rows_data.append((sort_key, row_values))

    # 降序排序：无值置底
    rows_data.sort(key=lambda x: (x[0] is None, -(x[0] if x[0] is not None else 0)))


    # 回写排序后的数据
    # 对 None 值写入空字符串（openpyxl 对已加载文件写 None 不会覆盖旧值）
    for i, (_, row_values) in enumerate(rows_data):
        target_row = i + 2
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=target_row, column=col_idx, value=value if value is not None else "")


    logging.info(f"Sheet '{ws.title}' 已按当日收益率降序排列")


# ===================== Sheet更新函数 =====================

def update_sheet(ws, name_map=None):
    """
    更新单个 Sheet：
    - 读取 A 列基金代码
    - 获取数据并写入相应列（A-J，包含波动率、夏普、成立以来收益率）
    - 按当日收益率降序排列（私募资管跳过）
    - 写入更新时间（K1）
    - 应用格式设置（不包含列宽调整）

    参数:
        ws: 工作表对象
        name_map: 全量基金名称映射字典，如果提供则优先使用
    返回:
        (总数量, 成功数量, 失败数量)
    """
    codes = get_codes_from_sheet(ws)
    total = len(codes)
    success = 0
    failed = 0

    if total == 0:
        write_update_time(ws)
        format_sheet(ws)
        logging.info(f"Sheet '{ws.title}' 无基金代码需要处理")
        return total, success, failed

    logging.info(f"开始处理 Sheet '{ws.title}'，共 {total} 个代码")

    for idx, (row_idx, code) in enumerate(codes, start=1):
        logging.info(f"[{ws.title}] {idx}/{total} 正在处理代码：{code}")
        try:
            data = fetch_fund_data(code)
            name = data.get("name")
            nav = data.get("nav")
            today_pct = data.get("today_pct")
            week_pct = data.get("week_pct")
            month_pct = data.get("month_pct")
            year_pct = data.get("year_pct")
            volatility_1y_pct = data.get("volatility_1y_pct")
            sharpe_1y = data.get("sharpe_1y")
            since_inception_pct = data.get("since_inception_pct")
            is_building_period = data.get("is_building_period", False)

            if is_building_period:
                logging.info(f"代码 {code} 为建仓期/封闭期基金，已取最新披露净值（非当日）或待手动补充")

            # 写入名称（B列）：优先使用name_map
            if name_map and code in name_map:
                name = name_map[code]
            ws.cell(row=row_idx, column=2, value=name if name else None)

            # 写入最新累计净值（C列）
            ws.cell(row=row_idx, column=3, value=nav if nav is not None else "")

            # 写入各收益率（D-G列）
            # 使用 "" 而非 None：openpyxl 对已加载文件写 None 不会覆盖旧值
            ws.cell(row=row_idx, column=4, value=(today_pct / 100.0 if today_pct is not None else ""))
            ws.cell(row=row_idx, column=5, value=(week_pct / 100.0 if week_pct is not None else ""))
            ws.cell(row=row_idx, column=6, value=(month_pct / 100.0 if month_pct is not None else ""))
            ws.cell(row=row_idx, column=7, value=(year_pct / 100.0 if year_pct is not None else ""))

            # 写入近1年波动率（H列，第8列）
            ws.cell(row=row_idx, column=8, value=(volatility_1y_pct / 100.0 if volatility_1y_pct is not None else ""))

            # 写入近1年夏普比率（I列，第9列）— 直接写数字，不除以100
            ws.cell(row=row_idx, column=9, value=sharpe_1y if sharpe_1y is not None else "")

            # 写入成立以来收益率（J列，第10列）
            ws.cell(row=row_idx, column=10, value=(since_inception_pct / 100.0 if since_inception_pct is not None else ""))

            success += 1
        except Exception as e:
            error_msg = str(e)
            # 优先从 name_map 获取名称
            name = None
            if name_map and code in name_map:
                name = name_map[code]
            else:
                try:
                    name = fetch_fund_name_from_akshare(code)
                except Exception:
                    pass

            # 检查是否是建仓期基金的异常
            if "建仓期" in error_msg or "封闭期" in error_msg:
                if name:
                    logging.info(f"代码 {code} 为建仓期/封闭期基金，已取最新披露净值（非当日）或待手动补充")
                    ws.cell(row=row_idx, column=2, value=name)
                    for col in range(3, 11):  # C到J列全部清空
                        ws.cell(row=row_idx, column=col, value="")
                    success += 1
                else:
                    logging.warning(f"[{ws.title}] 代码 {code} 数据获取失败：{e}")
                    ws.cell(row=row_idx, column=2, value="数据获取失败")
                    failed += 1
            else:
                if name:
                    logging.warning(f"[{ws.title}] 代码 {code} 净值获取失败，但已获取名称：{e}")
                    ws.cell(row=row_idx, column=2, value=name)
                    for col in range(3, 11):  # C到J列全部清空
                        ws.cell(row=row_idx, column=col, value="")
                    success += 1
                else:
                    logging.warning(f"[{ws.title}] 代码 {code} 数据获取失败：{e}")
                    ws.cell(row=row_idx, column=2, value="数据获取失败")
                    failed += 1
        finally:
            sleep_time = random.uniform(REQUEST_INTERVAL_MIN, REQUEST_INTERVAL_MAX)
            time.sleep(sleep_time)

    # 按当日收益率降序排列（私募资管Sheet自动跳过）
    sort_sheet_by_daily_return(ws)
    # 排列后重新应用格式设置
    format_sheet(ws)
    write_update_time(ws)
    logging.info(f"完成 Sheet '{ws.title}'：总数 {total}，成功 {success}，失败 {failed}")
    return total, success, failed



# ===================== 主函数 =====================

def main():
    """主函数：流程编排"""
    if not HAS_AKSHARE:
        logging.error("未检测到 akshare，无法获取基金数据。请先安装：pip install akshare")
        return
    
    # 显示启动信息
    logging.info("=" * 60)
    logging.info("基金净值更新器")
    logging.info(f"自动处理Sheet表头：{len(HEADERS)} 列（A-H列，包含成立以来收益率%）")
    logging.info("Excel格式：A列文本格式、自动调整列宽、数字列2位小数+%符号")
    avg_interval = (REQUEST_INTERVAL_MIN + REQUEST_INTERVAL_MAX) / 2
    logging.info(f"[INFO] 数据获取间隔设为 {avg_interval:.1f} 秒（随机范围：{REQUEST_INTERVAL_MIN}-{REQUEST_INTERVAL_MAX}秒）")
    logging.info("=" * 60)
    
    # 一次性获取全量基金名称映射
    logging.info("正在加载全量基金名称映射...")
    try:
        name_map = load_fund_name_map()
        fund_count = len(name_map)
        logging.info(f"[INFO] 已加载全量基金名称映射，共 {fund_count} 个基金")
    except Exception as e:
        logging.warning(f"加载全量基金名称映射失败：{e}，将使用备用方案")
        name_map = {}
    
    # 加载或创建当天的工作簿
    wb, file_path = load_or_inherit_workbook()
    
    # 更新自动处理的Sheet
    grand_total = 0
    grand_success = 0
    grand_failed = 0
    
    for sheet_name in AUTO_SHEETS:
        if sheet_name not in wb.sheetnames:
            logging.warning(f"工作簿中不存在 Sheet '{sheet_name}'，已跳过")
            continue
        ws = wb[sheet_name]
        total, success, failed = update_sheet(ws, name_map=name_map)
        grand_total += total
        grand_success += success
        grand_failed += failed
    
    # 在所有Sheet数据完全填充完毕后，统一调整列宽
    logging.info("正在调整所有Sheet的列宽...")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.title == MANUAL_SHEET:
            logging.info(f"[INFO] '{MANUAL_SHEET}' Sheet 保留手动内容，未进行任何自动处理")
        else:
            adjust_column_widths(ws)
    logging.info("[INFO] 列宽调整比例已优化（缩小约25%），长名称完整显示且不浪费空间")
    logging.info("列宽调整完成")
    
    # 保存前重新排序 Sheet，确保显示顺序正确
    reorder_sheets(wb)
    logging.info("Sheet 顺序已调整为：主推公募 → ETF → 私募资管 → 个人关注产品")
    
    # 保存前强制清除所有 None 单元格
    # openpyxl 对已加载文件的单元格设 value=None 后，XML 中仍可能保留旧值节点
    # 通过从 _cells 字典中彻底删除这些单元格，确保保存后 xlsx 中不含幽灵数据
    for sheet_name in AUTO_SHEETS:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            keys_to_delete = []
            for (row, col), cell in ws._cells.items():
                if row >= 2 and cell.value is None:
                    keys_to_delete.append((row, col))
            for key in keys_to_delete:
                del ws._cells[key]
            if keys_to_delete:
                logging.info(f"Sheet '{sheet_name}': 清除了 {len(keys_to_delete)} 个空单元格")
    
    # 为所有有数据的Sheet（包括私募资管）添加自动筛选和冻结窗格
    for sheet_name in wb.sheetnames:
        filter_ws = wb[sheet_name]
        # 只要有一行一列数据就加上筛选器
        if filter_ws.max_row >= 1 and filter_ws.max_column >= 1:
            filter_ws.auto_filter.ref = filter_ws.dimensions
            # 冻结第一行(表头)和第一列(代码)
            filter_ws.freeze_panes = "B2"
            logging.info(f"Sheet '{sheet_name}' 已启用表头自动筛选及首行首列冻结")

    # 保存文件（覆盖）
    wb.save(file_path)

    
    # 输出完成信息
    logging.info(f"更新完成：总共处理基金 {grand_total} 个，成功 {grand_success} 个，失败 {grand_failed} 个")
    logging.info("Excel格式已应用：A列文本格式、列宽已自动调整、数字格式（2位小数+%）")
    print(f"更新完成，处理了 {grand_total} 个基金（成功 {grand_success}，失败 {grand_failed}），文件已保存：{file_path}")


if __name__ == "__main__":
    main()
